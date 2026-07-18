#!/usr/bin/env python3
"""
gitlab_mr_monthly_report_pygitlab.py

Same report as gitlab_mr_monthly_report.py, but built on the `python-gitlab`
library instead of raw `requests` calls. Supports multiple projects and can
write a multi-tab Excel workbook (one sheet per project), and/or upsert the
results into a PostgreSQL table.

Install:
    pip install python-gitlab openpyxl psycopg2-binary

Usage examples
--------------
python gitlab_mr_monthly_report_pygitlab.py \
    --url https://gitlab.com \
    --project mygroup/myproject \
    --token $GITLAB_TOKEN

python gitlab_mr_monthly_report_pygitlab.py \
    --url https://gitlab.com \
    --project mygroup/proj1 mygroup/proj2 12345678 \
    --token $GITLAB_TOKEN \
    --group-by merged \
    --since 2025-01-01 \
    --until 2025-12-31 \
    --approvals \
    --xlsx mrs_2025.xlsx

# Upsert into PostgreSQL instead of (or alongside) file output
python gitlab_mr_monthly_report_pygitlab.py \
    --url https://gitlab.com \
    --project mygroup/proj1 mygroup/proj2 \
    --token $GITLAB_TOKEN \
    --approvals \
    --pg-dsn "postgresql://user:password@localhost:5432/mydb" \
    --pg-table gitlab_merge_requests \
    --no-print

# Or specify connection params individually instead of a DSN
python gitlab_mr_monthly_report_pygitlab.py \
    --url https://gitlab.com \
    --project mygroup/proj1 \
    --token $GITLAB_TOKEN \
    --pg-host localhost --pg-port 5432 --pg-dbname mydb --pg-user user --pg-password secret

Notes
-----
- `--project` accepts numeric project IDs and/or paths, given as
  multiple values and/or comma-separated within a single value, e.g.
  --project 123 groupA/proj1,groupA/proj2
- Each project gets its own sheet in the --xlsx workbook, named after
  the project's path (sanitized/truncated for Excel's naming rules,
  de-duplicated with a numeric suffix if needed).
- --csv writes a single combined CSV with a leading "project" column;
  --xlsx is what splits projects into separate tabs; --pg-* upserts into
  a single PostgreSQL table with a "project" column (all three can be
  used together in one run since they all read from the same fetched data).
- PostgreSQL output creates the target table if it doesn't exist, keyed
  on (project, iid), and upserts on conflict — safe to re-run on a
  schedule to keep the table current. Give either --pg-dsn (a full
  connection string) or the individual --pg-host/--pg-port/--pg-dbname/
  --pg-user/--pg-password flags.
- `--approvals` calls mr.approvals.get() per MR. On GitLab CE/Free this
  endpoint exists but approval rules are a paid feature, so approved_by
  may always be empty there ("-" will be shown, not an error).
- A personal/project access token with at least `read_api` scope is
  required for private projects.
"""

import argparse
import csv
import re
import sys
from collections import defaultdict
from datetime import datetime, timezone

try:
    import gitlab
except ImportError:
    sys.exit("This script requires python-gitlab: pip install python-gitlab")


FIELDNAMES = [
    "iid",
    "title",
    "state",
    "author",
    "created_at",
    "merged_at",
    "merged_by",
    "approved_by",
    "web_url",
]


def parse_args():
    p = argparse.ArgumentParser(
        description="Report GitLab merge requests across one or more projects, grouped by month, "
        "with status and merge/approval info (python-gitlab version)."
    )
    p.add_argument("--url", required=True, help="Base GitLab URL, e.g. https://gitlab.com")
    p.add_argument(
        "--project",
        required=True,
        nargs="+",
        help="One or more project IDs and/or paths (e.g. group/subgroup/project). "
        "Can be repeated and/or comma-separated, e.g. --project 123 groupA/proj1,groupA/proj2",
    )
    p.add_argument("--token", required=True, help="Personal/project access token (needs read_api scope)")
    p.add_argument(
        "--state",
        default="all",
        choices=["all", "opened", "closed", "merged", "locked"],
        help="Filter MRs by state (default: all)",
    )
    p.add_argument(
        "--group-by",
        default="created",
        choices=["created", "merged", "updated"],
        help="Which date field to group MRs by month on (default: created)",
    )
    p.add_argument("--since", help="Only include MRs on/after this date (YYYY-MM-DD), applied to the group-by field")
    p.add_argument("--until", help="Only include MRs on/before this date (YYYY-MM-DD), applied to the group-by field")
    p.add_argument(
        "--approvals",
        action="store_true",
        help="Fetch approval info per MR (extra API call per MR — slower, but shows who approved it)",
    )
    p.add_argument("--csv", metavar="FILE", help="Write a single combined CSV (all projects, with a 'project' column)")
    p.add_argument("--xlsx", metavar="FILE", help="Write an Excel workbook with one sheet per project")
    p.add_argument("--no-print", action="store_true", help="Suppress console table output (useful with --csv/--xlsx)")
    p.add_argument("--insecure", action="store_true", help="Disable TLS certificate verification")

    pg_group = p.add_argument_group("PostgreSQL output")
    pg_group.add_argument(
        "--pg-dsn",
        help="Full PostgreSQL connection string, e.g. postgresql://user:pass@host:5432/dbname. "
        "Takes precedence over the individual --pg-* flags below if given.",
    )
    pg_group.add_argument("--pg-host", help="PostgreSQL host (alternative to --pg-dsn)")
    pg_group.add_argument("--pg-port", type=int, default=5432, help="PostgreSQL port (default 5432)")
    pg_group.add_argument("--pg-dbname", help="PostgreSQL database name")
    pg_group.add_argument("--pg-user", help="PostgreSQL user")
    pg_group.add_argument("--pg-password", help="PostgreSQL password")
    pg_group.add_argument(
        "--pg-table",
        default="gitlab_merge_requests",
        help="Target table name (created if missing, default: gitlab_merge_requests)",
    )
    pg_group.add_argument("--pg-schema", help="Optional schema name to qualify the table with")

    return p.parse_args()


def expand_projects(raw_list):
    """Flatten repeated --project values and comma-separated entries, de-duplicated, order preserved."""
    seen = set()
    projects = []
    for raw in raw_list:
        for item in raw.split(","):
            item = item.strip()
            if item and item not in seen:
                seen.add(item)
                projects.append(item)
    return projects


def month_key(iso_str):
    if not iso_str:
        return None
    dt = datetime.strptime(iso_str[:10], "%Y-%m-%d")
    return dt.strftime("%Y-%m")


def in_range(iso_str, since, until):
    if not iso_str:
        return since is None and until is None
    dt = datetime.strptime(iso_str[:10], "%Y-%m-%d")
    if since and dt < since:
        return False
    if until and dt > until:
        return False
    return True


def sanitize_sheet_name(name, existing):
    """Excel sheet names: max 31 chars, no [ ] : * ? / \\, not blank, must be unique."""
    cleaned = re.sub(r"[\[\]:*?/\\]", "_", name).strip() or "project"
    cleaned = cleaned[:31]
    candidate = cleaned
    suffix = 1
    while candidate.lower() in existing:
        tail = f"_{suffix}"
        candidate = cleaned[: 31 - len(tail)] + tail
        suffix += 1
    existing.add(candidate.lower())
    return candidate


def fetch_project_rows(gl, project_identifier, args, since, until):
    """Fetch and shape MR rows for a single project. Returns (display_name, rows)."""
    project = gl.projects.get(project_identifier)
    display_name = project.path_with_namespace

    list_kwargs = {"order_by": "created_at", "sort": "asc", "all": True}
    if args.state != "all":
        list_kwargs["state"] = args.state

    mrs = project.mergerequests.list(**list_kwargs)

    date_field_map = {"created": "created_at", "merged": "merged_at", "updated": "updated_at"}
    date_field = date_field_map[args.group_by]

    rows = []
    for mr in mrs:
        group_date = getattr(mr, date_field, None)

        if args.group_by == "merged" and not group_date:
            continue
        if not in_range(group_date, since, until):
            continue

        approvers = "-"
        if args.approvals:
            try:
                approval_info = mr.approvals.get()
                approved_by = getattr(approval_info, "approved_by", None) or []
                approvers = ", ".join(a["user"]["username"] for a in approved_by) if approved_by else "-"
            except gitlab.exceptions.GitlabGetError:
                approvers = "n/a"

        merged_by = getattr(mr, "merged_by", None)
        merged_by_username = merged_by["username"] if merged_by else "-"

        rows.append(
            {
                "month": month_key(group_date) or "(no date)",
                "iid": mr.iid,
                "title": mr.title,
                "state": mr.state,
                "author": mr.author["username"] if getattr(mr, "author", None) else "-",
                "created_at": (mr.created_at or "")[:10],
                "merged_at": (mr.merged_at or "-")[:10] if getattr(mr, "merged_at", None) else "-",
                "merged_by": merged_by_username,
                "approved_by": approvers,
                "web_url": mr.web_url,
            }
        )

    return display_name, rows


def print_console(display_name, rows, show_approvals):
    grouped = defaultdict(list)
    for row in rows:
        grouped[row["month"]].append(row)

    print(f"\n########## Project: {display_name} ##########")
    for month in sorted(grouped.keys()):
        month_rows = grouped[month]
        print(f"\n=== {month} ({len(month_rows)} MR{'s' if len(month_rows) != 1 else ''}) ===")
        for r in month_rows:
            print(f"  !{r['iid']:<5} [{r['state']:<8}] {r['title'][:60]:<60}")
            print(f"          author: {r['author']:<15} created: {r['created_at']}")
            if r["state"] == "merged":
                print(f"          merged_at: {r['merged_at']:<12} merged_by: {r['merged_by']}")
                if show_approvals:
                    print(f"          approved_by: {r['approved_by']}")
            print(f"          {r['web_url']}")

    total = len(rows)
    merged_count = sum(1 for r in rows if r["state"] == "merged")
    print(f"\nTotal MRs: {total} | Merged: {merged_count} | Months: {len(grouped)}")


def write_combined_csv(path, project_rows):
    fieldnames = ["project", "month"] + FIELDNAMES
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for display_name, rows in project_rows:
            grouped = defaultdict(list)
            for row in rows:
                grouped[row["month"]].append(row)
            for month in sorted(grouped.keys()):
                for r in grouped[month]:
                    out = {"project": display_name, "month": month}
                    out.update({k: r[k] for k in FIELDNAMES})
                    writer.writerow(out)
    print(f"\nCSV written to {path}", file=sys.stderr)


def write_xlsx(path, project_rows):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter
    except ImportError:
        sys.exit("Writing --xlsx requires openpyxl: pip install openpyxl")

    headers = ["Month"] + FIELDNAMES

    wb = Workbook()
    wb.remove(wb.active)
    used_names = set()

    for display_name, rows in project_rows:
        sheet_name = sanitize_sheet_name(display_name, used_names)
        ws = wb.create_sheet(title=sheet_name)

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(name="Arial", bold=True)
        ws.freeze_panes = "A2"

        grouped = defaultdict(list)
        for row in rows:
            grouped[row["month"]].append(row)

        row_idx = 2
        col_widths = [len(h) for h in headers]
        for month in sorted(grouped.keys()):
            for r in grouped[month]:
                values = [month] + [r[k] for k in FIELDNAMES]
                for col_idx, value in enumerate(values, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.font = Font(name="Arial")
                    col_widths[col_idx - 1] = max(col_widths[col_idx - 1], len(str(value)))
                row_idx += 1

        for col_idx, width in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(width + 2, 10), 60)

        if not rows:
            ws.cell(row=2, column=1, value="(no merge requests matched the given filters)").font = Font(
                name="Arial", italic=True
            )

    wb.save(path)
    print(f"\nExcel workbook written to {path} ({len(project_rows)} sheet(s))", file=sys.stderr)


def build_pg_conn_kwargs(args):
    """Build psycopg2.connect() kwargs from --pg-dsn or the individual --pg-* flags."""
    if args.pg_dsn:
        return {"dsn": args.pg_dsn}
    if not all([args.pg_host, args.pg_dbname, args.pg_user]):
        sys.exit(
            "PostgreSQL output requires either --pg-dsn, or at minimum "
            "--pg-host, --pg-dbname, and --pg-user (plus --pg-password if needed)."
        )
    return {
        "host": args.pg_host,
        "port": args.pg_port,
        "dbname": args.pg_dbname,
        "user": args.pg_user,
        "password": args.pg_password,
    }


def _to_date_or_none(value):
    """Convert a 'YYYY-MM-DD' string to a date, or None for placeholders like '-' / ''."""
    if not value or value == "-":
        return None
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except ValueError:
        return None


def _to_text_or_none(value):
    return None if (value is None or value == "-") else value


def write_postgres(args, project_rows):
    try:
        import psycopg2
        from psycopg2.extras import execute_values
    except ImportError:
        sys.exit("PostgreSQL output requires psycopg2: pip install psycopg2-binary")

    conn_kwargs = build_pg_conn_kwargs(args)
    table_ident = f'"{args.pg_schema}"."{args.pg_table}"' if args.pg_schema else f'"{args.pg_table}"'

    create_sql = f"""
        CREATE TABLE IF NOT EXISTS {table_ident} (
            project      TEXT NOT NULL,
            month        TEXT NOT NULL,
            iid          INTEGER NOT NULL,
            title        TEXT,
            state        TEXT,
            author       TEXT,
            created_at   DATE,
            merged_at    DATE,
            merged_by    TEXT,
            approved_by  TEXT,
            web_url      TEXT,
            synced_at    TIMESTAMPTZ NOT NULL DEFAULT now(),
            PRIMARY KEY (project, iid)
        );
    """

    upsert_sql = f"""
        INSERT INTO {table_ident}
            (project, month, iid, title, state, author, created_at, merged_at,
             merged_by, approved_by, web_url, synced_at)
        VALUES %s
        ON CONFLICT (project, iid) DO UPDATE SET
            month       = EXCLUDED.month,
            title       = EXCLUDED.title,
            state       = EXCLUDED.state,
            author      = EXCLUDED.author,
            created_at  = EXCLUDED.created_at,
            merged_at   = EXCLUDED.merged_at,
            merged_by   = EXCLUDED.merged_by,
            approved_by = EXCLUDED.approved_by,
            web_url     = EXCLUDED.web_url,
            synced_at   = EXCLUDED.synced_at;
    """

    now = datetime.now(timezone.utc)
    total_rows = 0

    conn = psycopg2.connect(**conn_kwargs)
    try:
        with conn.cursor() as cur:
            cur.execute(create_sql)
        conn.commit()

        for display_name, rows in project_rows:
            if not rows:
                continue
            values = [
                (
                    display_name,
                    r["month"],
                    r["iid"],
                    r["title"],
                    r["state"],
                    _to_text_or_none(r["author"]),
                    _to_date_or_none(r["created_at"]),
                    _to_date_or_none(r["merged_at"]),
                    _to_text_or_none(r["merged_by"]),
                    _to_text_or_none(r["approved_by"]),
                    r["web_url"],
                    now,
                )
                for r in rows
            ]
            with conn.cursor() as cur:
                execute_values(cur, upsert_sql, values)
            conn.commit()
            total_rows += len(values)
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()

    print(
        f"\nPostgreSQL: upserted {total_rows} row(s) across {len(project_rows)} project(s) into {table_ident}",
        file=sys.stderr,
    )


def main():
    args = parse_args()
    projects = expand_projects(args.project)

    since = datetime.strptime(args.since, "%Y-%m-%d") if args.since else None
    until = datetime.strptime(args.until, "%Y-%m-%d") if args.until else None

    gl = gitlab.Gitlab(args.url, private_token=args.token, ssl_verify=not args.insecure)
    try:
        gl.auth()
    except gitlab.exceptions.GitlabAuthenticationError:
        sys.exit("Authentication failed. Check your --token.")

    project_rows = []
    failures = []
    for identifier in projects:
        print(f"Fetching merge requests for project '{identifier}' (state={args.state})...", file=sys.stderr)
        try:
            display_name, rows = fetch_project_rows(gl, identifier, args, since, until)
        except gitlab.exceptions.GitlabGetError as e:
            print(f"  Skipping '{identifier}': {e}", file=sys.stderr)
            failures.append(identifier)
            continue
        project_rows.append((display_name, rows))

    if not project_rows:
        sys.exit("No project data was fetched successfully. Check --project values, --token, and --url.")

    if not args.no_print:
        for display_name, rows in project_rows:
            print_console(display_name, rows, args.approvals)

    if args.csv:
        write_combined_csv(args.csv, project_rows)

    if args.xlsx:
        write_xlsx(args.xlsx, project_rows)

    if args.pg_dsn or args.pg_host:
        write_postgres(args, project_rows)

    if failures:
        print(f"\nNote: {len(failures)} project(s) were skipped due to errors: {', '.join(failures)}", file=sys.stderr)


if __name__ == "__main__":
    main()
