#!/usr/bin/env python3
"""
gitlab_mr_monthly_report.py

List all merge requests across one or more GitLab projects, grouped by
month, showing status information per MR and — for merged MRs — who
merged/approved it. Can write a multi-tab Excel workbook (one sheet per
project) in addition to console/CSV output.

Requires:
    pip install requests openpyxl

Usage examples
--------------
# Single project, printed to console (original behaviour still works)
python gitlab_mr_monthly_report.py --url https://gitlab.com --project 12345678 --token $GITLAB_TOKEN

# Multiple projects (mix of numeric IDs and paths), one tab per project in Excel
python gitlab_mr_monthly_report.py \
    --url https://gitlab.com \
    --project mygroup/myproject 12345678 anothergroup/otherproject \
    --token $GITLAB_TOKEN \
    --group-by merged \
    --approvals \
    --xlsx mr_report.xlsx

# Comma-separated also works, and can be combined with repeated --project flags
python gitlab_mr_monthly_report.py \
    --url https://gitlab.com \
    --project "group1/proj1,group1/proj2" --project 987654 \
    --token $GITLAB_TOKEN \
    --xlsx mr_report.xlsx \
    --no-print

Notes
-----
- `--project` accepts numeric project IDs and/or URL-encodable paths
  (e.g. "mygroup/myproject"), any mix of the two, given as multiple
  values and/or comma-separated within a single value.
- Each project gets its own sheet in the --xlsx workbook, named after
  the project's path (truncated/sanitized to fit Excel's 31-character,
  no-special-character sheet name limit). If two projects sanitize to
  the same name, a numeric suffix is added to keep them unique.
- --csv now writes a single combined CSV with a leading "project"
  column (useful for pivoting across all projects at once); --xlsx is
  the option that splits projects into separate tabs.
- Approval information (`--approvals`) uses the
  /merge_requests/:iid/approvals endpoint. On GitLab CE/Free instances
  this endpoint exists but approval *rules* are a paid feature, so the
  "approved_by" list may always be empty there — the script won't error,
  it will just show "-" for approvers on those instances.
- A personal/project access token with at least `read_api` scope is
  required for private projects.
"""

import argparse
import csv
import re
import sys
from collections import defaultdict
from datetime import datetime
from urllib.parse import quote

try:
    import requests
except ImportError:
    sys.exit("This script requires the 'requests' package: pip install requests")


FIELDNAMES = [
    "iid",
    "title",
    "state",
    "author",
    "created_at",
    "merged_at",
    "merged_by",
    "approved_by",
    "web_url",
]


def parse_args():
    p = argparse.ArgumentParser(
        description="Report GitLab merge requests across one or more projects, grouped by month, "
        "with status and merge/approval info."
    )
    p.add_argument("--url", required=True, help="Base GitLab URL, e.g. https://gitlab.com")
    p.add_argument(
        "--project",
        required=True,
        nargs="+",
        help="One or more project IDs and/or paths (e.g. group/subgroup/project). "
        "Can be repeated and/or comma-separated, e.g. --project 123 groupA/proj1,groupA/proj2",
    )
    p.add_argument("--token", required=True, help="Personal/project access token (needs read_api scope)")
    p.add_argument(
        "--state",
        default="all",
        choices=["all", "opened", "closed", "merged", "locked"],
        help="Filter MRs by state (default: all)",
    )
    p.add_argument(
        "--group-by",
        default="created",
        choices=["created", "merged", "updated"],
        help="Which date field to group MRs by month on (default: created)",
    )
    p.add_argument("--since", help="Only include MRs on/after this date (YYYY-MM-DD), applied to the group-by field")
    p.add_argument("--until", help="Only include MRs on/before this date (YYYY-MM-DD), applied to the group-by field")
    p.add_argument(
        "--approvals",
        action="store_true",
        help="Fetch approval info per MR (extra API call per MR — slower, but shows who approved it)",
    )
    p.add_argument("--csv", metavar="FILE", help="Write a single combined CSV (all projects, with a 'project' column)")
    p.add_argument("--xlsx", metavar="FILE", help="Write an Excel workbook with one sheet per project")
    p.add_argument("--no-print", action="store_true", help="Suppress console table output (useful with --csv/--xlsx)")
    p.add_argument("--per-page", type=int, default=100, help="API page size (default 100, max 100)")
    p.add_argument("--insecure", action="store_true", help="Disable TLS certificate verification")
    return p.parse_args()


def expand_projects(raw_list):
    """Flatten repeated --project values and comma-separated entries, de-duplicated, order preserved."""
    seen = set()
    projects = []
    for raw in raw_list:
        for item in raw.split(","):
            item = item.strip()
            if item and item not in seen:
                seen.add(item)
                projects.append(item)
    return projects


def encode_project(identifier):
    return identifier if identifier.isdigit() else quote(identifier, safe="")


def api_get(session, base_url, path, params=None):
    """GET with pagination handling. Returns a list of all items across pages."""
    url = f"{base_url}/api/v4{path}"
    results = []
    params = dict(params or {})
    params["page"] = 1
    while True:
        resp = session.get(url, params=params, timeout=30)
        if resp.status_code == 401:
            sys.exit("Authentication failed (401). Check your --token.")
        if resp.status_code == 403:
            raise PermissionError(f"Forbidden (403) for {url}")
        if resp.status_code == 404:
            raise LookupError(f"Not found (404) for {url}")
        resp.raise_for_status()
        data = resp.json()
        if isinstance(data, list):
            results.extend(data)
        else:
            return data

        next_page = resp.headers.get("X-Next-Page")
        if not next_page:
            break
        params["page"] = next_page
    return results


def single_get(session, base_url, path, params=None):
    """GET a single (non-paginated) resource. Returns None on 404/403."""
    url = f"{base_url}/api/v4{path}"
    resp = session.get(url, params=params or {}, timeout=30)
    if resp.status_code in (403, 404):
        return None
    resp.raise_for_status()
    return resp.json()


def month_key(iso_str):
    if not iso_str:
        return None
    dt = datetime.strptime(iso_str[:10], "%Y-%m-%d")
    return dt.strftime("%Y-%m")


def in_range(iso_str, since, until):
    if not iso_str:
        return since is None and until is None
    dt = datetime.strptime(iso_str[:10], "%Y-%m-%d")
    if since and dt < since:
        return False
    if until and dt > until:
        return False
    return True


def sanitize_sheet_name(name, existing):
    """Excel sheet names: max 31 chars, no [ ] : * ? / \\, not blank, must be unique."""
    cleaned = re.sub(r"[\[\]:*?/\\]", "_", name).strip() or "project"
    cleaned = cleaned[:31]
    candidate = cleaned
    suffix = 1
    while candidate.lower() in existing:
        tail = f"_{suffix}"
        candidate = cleaned[: 31 - len(tail)] + tail
        suffix += 1
    existing.add(candidate.lower())
    return candidate


def fetch_project_rows(session, base_url, project_identifier, args, since, until):
    """Fetch and shape MR rows for a single project. Returns (display_name, rows)."""
    project = encode_project(project_identifier)
    proj_info = single_get(session, base_url, f"/projects/{project}")
    display_name = proj_info["path_with_namespace"] if proj_info else str(project_identifier)

    mrs = api_get(
        session,
        base_url,
        f"/projects/{project}/merge_requests",
        params={"state": args.state, "per_page": args.per_page, "order_by": "created_at", "sort": "asc"},
    )

    date_field_map = {"created": "created_at", "merged": "merged_at", "updated": "updated_at"}
    date_field = date_field_map[args.group_by]

    rows = []
    for mr in mrs:
        group_date = mr.get(date_field)

        if args.group_by == "merged" and not group_date:
            continue
        if not in_range(group_date, since, until):
            continue

        approvers = "-"
        if args.approvals:
            approvals_data = single_get(
                session, base_url, f"/projects/{project}/merge_requests/{mr['iid']}/approvals"
            )
            if approvals_data and approvals_data.get("approved_by"):
                names = [a["user"]["username"] for a in approvals_data["approved_by"]]
                approvers = ", ".join(names) if names else "-"
            elif approvals_data is not None:
                approvers = "-"
            else:
                approvers = "n/a"

        merged_by = mr.get("merged_by") or {}
        merged_by_username = merged_by.get("username", "-") if merged_by else "-"

        rows.append(
            {
                "month": month_key(group_date) or "(no date)",
                "iid": mr["iid"],
                "title": mr["title"],
                "state": mr["state"],
                "author": mr["author"]["username"] if mr.get("author") else "-",
                "created_at": (mr.get("created_at") or "")[:10],
                "merged_at": (mr.get("merged_at") or "-")[:10] if mr.get("merged_at") else "-",
                "merged_by": merged_by_username,
                "approved_by": approvers,
                "web_url": mr.get("web_url", ""),
            }
        )

    return display_name, rows


def print_console(display_name, rows, show_approvals):
    grouped = defaultdict(list)
    for row in rows:
        grouped[row["month"]].append(row)

    print(f"\n########## Project: {display_name} ##########")
    for month in sorted(grouped.keys()):
        month_rows = grouped[month]
        print(f"\n=== {month} ({len(month_rows)} MR{'s' if len(month_rows) != 1 else ''}) ===")
        for r in month_rows:
            print(f"  !{r['iid']:<5} [{r['state']:<8}] {r['title'][:60]:<60}")
            print(f"          author: {r['author']:<15} created: {r['created_at']}")
            if r["state"] == "merged":
                print(f"          merged_at: {r['merged_at']:<12} merged_by: {r['merged_by']}")
                if show_approvals:
                    print(f"          approved_by: {r['approved_by']}")
            print(f"          {r['web_url']}")

    total = len(rows)
    merged_count = sum(1 for r in rows if r["state"] == "merged")
    print(f"\nTotal MRs: {total} | Merged: {merged_count} | Months: {len(grouped)}")


def write_combined_csv(path, project_rows):
    fieldnames = ["project", "month"] + FIELDNAMES
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for display_name, rows in project_rows:
            grouped = defaultdict(list)
            for row in rows:
                grouped[row["month"]].append(row)
            for month in sorted(grouped.keys()):
                for r in grouped[month]:
                    out = {"project": display_name, "month": month}
                    out.update({k: r[k] for k in FIELDNAMES})
                    writer.writerow(out)
    print(f"\nCSV written to {path}", file=sys.stderr)


def write_xlsx(path, project_rows):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter
    except ImportError:
        sys.exit("Writing --xlsx requires openpyxl: pip install openpyxl")

    headers = ["Month"] + FIELDNAMES

    wb = Workbook()
    wb.remove(wb.active)  # drop the default blank sheet
    used_names = set()

    for display_name, rows in project_rows:
        sheet_name = sanitize_sheet_name(display_name, used_names)
        ws = wb.create_sheet(title=sheet_name)

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(name="Arial", bold=True)
        ws.freeze_panes = "A2"

        grouped = defaultdict(list)
        for row in rows:
            grouped[row["month"]].append(row)

        row_idx = 2
        col_widths = [len(h) for h in headers]
        for month in sorted(grouped.keys()):
            for r in grouped[month]:
                values = [month] + [r[k] for k in FIELDNAMES]
                for col_idx, value in enumerate(values, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.font = Font(name="Arial")
                    col_widths[col_idx - 1] = max(col_widths[col_idx - 1], len(str(value)))
                row_idx += 1

        for col_idx, width in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(width + 2, 10), 60)

        if not rows:
            ws.cell(row=2, column=1, value="(no merge requests matched the given filters)").font = Font(
                name="Arial", italic=True
            )

    wb.save(path)
    print(f"\nExcel workbook written to {path} ({len(project_rows)} sheet(s))", file=sys.stderr)


def main():
    args = parse_args()
    base_url = args.url.rstrip("/")
    projects = expand_projects(args.project)

    since = datetime.strptime(args.since, "%Y-%m-%d") if args.since else None
    until = datetime.strptime(args.until, "%Y-%m-%d") if args.until else None

    session = requests.Session()
    session.headers.update({"PRIVATE-TOKEN": args.token})
    if args.insecure:
        session.verify = False
        import urllib3

        urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

    project_rows = []
    failures = []
    for identifier in projects:
        print(f"Fetching merge requests for project '{identifier}' (state={args.state})...", file=sys.stderr)
        try:
            display_name, rows = fetch_project_rows(session, base_url, identifier, args, since, until)
        except (PermissionError, LookupError) as e:
            print(f"  Skipping '{identifier}': {e}", file=sys.stderr)
            failures.append(identifier)
            continue
        project_rows.append((display_name, rows))

    if not project_rows:
        sys.exit("No project data was fetched successfully. Check --project values, --token, and --url.")

    if not args.no_print:
        for display_name, rows in project_rows:
            print_console(display_name, rows, args.approvals)

    if args.csv:
        write_combined_csv(args.csv, project_rows)

    if args.xlsx:
        write_xlsx(args.xlsx, project_rows)

    if failures:
        print(f"\nNote: {len(failures)} project(s) were skipped due to errors: {', '.join(failures)}", file=sys.stderr)


if __name__ == "__main__":
    main()
